import os
import cv2
import openpyxl
from openpyxl import Workbook, load_workbook
from tkinter import Tk, Label, Button, Checkbutton, IntVar, Frame, StringVar, Canvas, Toplevel, Listbox, Scrollbar, messagebox, Scale, simpledialog, Entry
from tkinter import ttk
from PIL import Image, ImageTk, ImageDraw, ImageEnhance
from datetime import datetime
import re
import random
import json
import sys
import csv
import traceback
import glob

# --- NEW LOGIN DIALOGS ---

def prompt_for_username():
    """Creates a simple dialog to get just the username."""
    dialog = Tk()
    dialog.title("Login")
    
    username_var = StringVar()

    def on_login():
        if entry.get().strip():
            username_var.set(entry.get().strip())
            dialog.destroy()

    def on_cancel():
        username_var.set("")
        dialog.destroy()

    dialog.protocol("WM_DELETE_WINDOW", on_cancel)
    dialog.bind('<Return>', lambda event: on_login())

    main_frame = Frame(dialog, padx=30, pady=20)
    main_frame.pack()

    Label(main_frame, text="Please enter your name:", font=("Arial", 16)).pack(pady=(0, 10))
    
    entry = Entry(main_frame, font=("Arial", 16), width=40)
    entry.pack(pady=10)
    entry.focus_set()

    button_frame = Frame(main_frame)
    button_frame.pack(pady=20)

    Button(button_frame, text="Login", command=on_login, font=("Arial", 14), width=12).pack(side="left", padx=10)
    Button(button_frame, text="Cancel", command=on_cancel, font=("Arial", 14), width=12).pack(side="right", padx=10)

    dialog.update_idletasks()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    x = (dialog.winfo_screenwidth() // 2) - (width // 2)
    y = (dialog.winfo_screenheight() // 2) - (height // 2)
    dialog.geometry(f'{width}x{height}+{x}+{y}')
    dialog.resizable(False, False)
    
    dialog.grab_set()
    dialog.wait_window()
    
    return username_var.get()

def prompt_for_expertise():
    """Creates a separate dialog to get expertise for a new user. Returns a score."""
    dialog = Tk()
    dialog.title("New User Setup")
    
    expertise_var = StringVar()
    # This will hold the integer score
    score_to_return = None

    expertise_to_score = {
        "No archaeology experience": 0,
        "Archaeology experience but no Remote Sensing experience": 1,
        "Some archaeology and Remote Sensing experience": 3,
        "Remote Sensing specialist": 5,
        "Prehistoric enclosures specialist": 5
    }

    def on_confirm():
        nonlocal score_to_return
        selected_option = expertise_var.get()
        if selected_option:
            score_to_return = expertise_to_score.get(selected_option)
            dialog.destroy()

    def on_cancel():
        nonlocal score_to_return
        score_to_return = None
        dialog.destroy()

    dialog.protocol("WM_DELETE_WINDOW", on_cancel)
    dialog.bind('<Return>', lambda event: on_confirm())

    style = ttk.Style(dialog)
    style.configure('Custom.TRadiobutton', font=('Arial', 16))

    main_frame = Frame(dialog, padx=30, pady=20)
    main_frame.pack()

    Label(main_frame, text="New user detected. Which best describes you?", font=("Arial", 16)).pack(anchor='w', pady=(0, 10))

    # Use the keys from the dictionary as options
    for option in expertise_to_score.keys():
        ttk.Radiobutton(main_frame, text=option, variable=expertise_var, value=option, style='Custom.TRadiobutton').pack(anchor='w', padx=10, pady=2)

    button_frame = Frame(main_frame)
    button_frame.pack(pady=20)

    Button(button_frame, text="Confirm", command=on_confirm, font=("Arial", 14), width=12).pack()

    dialog.update_idletasks()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    x = (dialog.winfo_screenwidth() // 2) - (width // 2)
    y = (dialog.winfo_screenheight() // 2) - (height // 2)
    dialog.geometry(f'{width}x{height}+{x}+{y}')
    dialog.resizable(False, False)
    
    dialog.grab_set()
    dialog.wait_window()
    
    return score_to_return

# --- PyInstaller-aware path configuration ---
# Determine the base directory, whether running as a script or a frozen exe
if getattr(sys, 'frozen', False):
    # If the application is run as a bundle, the script_dir is where the exe is
    script_dir = os.path.dirname(sys.executable)
else:
    # If running as a normal script, the script_dir is the script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

# Path to the main folder, now located in the same directory as the script/exe
main_folder = os.path.join(script_dir, 'image_dataset')

# Define the path for the annotations folder and create it if it doesn't exist
annotations_folder = os.path.join(script_dir, 'annotations')
os.makedirs(annotations_folder, exist_ok=True)

def log(msg: str):
    try:
        with open(os.path.join(annotations_folder, "app.log"), "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except Exception:
        pass

# --- NEW LOGIN WORKFLOW ---
username = prompt_for_username()

if not username:
    print("Login cancelled. Exiting.")
    sys.exit()

# Search for an existing file for the user
existing_files = glob.glob(os.path.join(annotations_folder, f"cropmarks_{username}_*.xlsx"))

if existing_files:
    # User exists, use the first file found
    excel_path = existing_files[0]
    print(f"Welcome back, {username}. Loading session from: {excel_path}")
else:
    # New user, prompt for expertise to get the score
    score = prompt_for_expertise()
    if score is not None:
        excel_path = os.path.join(annotations_folder, f"cropmarks_{username}_{score}.xlsx")
        print(f"New user '{username}' created. Data will be saved to: {excel_path}")
    else:
        print("New user setup cancelled. Exiting.")
        sys.exit()

# --- End of User Setup ---


# Initialize the workbook and worksheet
is_new_file = not os.path.exists(excel_path)
if not is_new_file:
    wb = load_workbook(excel_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"
    # Add the new 'Order' column
    ws.append(['Site', 'Date', 'Cropmark', 'Image', 'Saved', 'Drawing', 'QC_Reference', 'UniqueID', 'Order'])

# Function to save the workbook
def save_workbook():
    wb.save(excel_path)

# Get the list of folders
folders = [os.path.join(main_folder, folder) for folder in os.listdir(main_folder) if os.path.isdir(os.path.join(main_folder, folder))]

# Create a single list of all unique image paths first
unique_image_paths = []
for folder in folders:
    images_in_folder = [os.path.join(folder, img) for img in os.listdir(folder) if img.lower().endswith('.jpg')]
    unique_image_paths.extend(images_in_folder)

# --- Session Restore or New Session Creation ---
all_images = []
initial_image_idx = 0

if not is_new_file:
    print("Loading existing session from Excel file...")
    # Read all data from the Excel file into a temporary list
    temp_image_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Assuming columns are: Site, Date, Cropmark, Image, Saved, Drawing, QC_Ref, UniqueID, Order
        if len(row) >= 9 and row[7] is not None: # Check for UniqueID
             # Find the full path from the unique_image_paths list
            image_name = row[3]
            full_path = next((p for p in unique_image_paths if os.path.basename(p) == image_name), None)
            if full_path:
                temp_image_list.append({
                    'path': full_path,
                    'qc_ref': row[6],
                    'id': row[7],
                    'order': row[8],
                    'saved': row[4] == 'S'
                })

    # Sort the list based on the saved order
    temp_image_list.sort(key=lambda x: x['order'])
    all_images = temp_image_list

    # Find the first unsaved image
    try:
        initial_image_idx = next(i for i, img in enumerate(all_images) if not img['saved'])
    except StopIteration:
        print("All images are already saved!")
        initial_image_idx = 0 # Default to first image if all are done

    print(f"Session restored. Starting at image index {initial_image_idx}.")

else: # This is a new user/file
    print("Creating a new session...")
    # 1. Create a list of dictionaries for the original images.
    #    The unique ID for an original is its filename.
    temp_image_list = []
    for path in unique_image_paths:
        filename = os.path.basename(path)
        temp_image_list.append({'path': path, 'qc_ref': None, 'id': filename})

    # 2. Select 39 random but DETERMINISTIC images to be used for QC.
    num_duplicates = 39
    if len(unique_image_paths) > 0:
        # Sort the paths to ensure the list is in the same order every time
        unique_image_paths.sort()
        
        # Use a fixed seed to make the random selection deterministic
        random.seed(42) # The number 42 is arbitrary; any constant number works.
        
        # Ensure we don't try to sample more images than exist
        num_to_sample = min(num_duplicates, len(unique_image_paths))
        
        qc_images_to_duplicate = random.sample(unique_image_paths, num_to_sample)
        
        # Reset the seed so the final shuffle is different each time
        random.seed()

        # 3. Create the duplicate entries with a stable, unique QC ID
        for qc_path in qc_images_to_duplicate:
            qc_filename = os.path.basename(qc_path)
            # The duplicate entry points back to the original's filename
            # The ID is made unique by adding a suffix.
            temp_image_list.append({'path': qc_path, 'qc_ref': qc_filename, 'id': f"{qc_filename}_qc"})

    # 4. Shuffle the entire list to randomize the presentation order for the user
    random.shuffle(temp_image_list)
    
    # 5. Assign the final order index to the shuffled list
    for i, img_data in enumerate(temp_image_list):
        img_data['order'] = i
    
    all_images = temp_image_list
# --- End of Session Logic ---


# Mapping of months to numeric values
month_map = {
    'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04', 'mai': '05', 'jun': '06',
    'jul': '07', 'ago': '08', 'set': '09', 'out': '10', 'nov': '11', 'dez': '12',
    'apr': '04', 'aug': '08', 'dec': '12', 'feb': '02', 'jan': '01', 'jul': '07',
    'jun': '06', 'mar': '03', 'may': '05', 'nov': '11', 'oct': '10', 'sep': '09'
}

# Function to convert month to a numeric value
def month_to_number(month_str):
    return month_map.get(month_str.lower(), '00')

# --- Pre-populate Excel on startup ---
def populate_excel_on_startup():
    """Checks for and adds any new images to the Excel file on startup."""
    print("Checking for new images to add to the Excel file...")
    
    # Get a set of all unique IDs already in the Excel file for fast checking
    existing_ids = set()
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, values_only=True):
            if row[0] is not None:
                existing_ids.add(row[0])

    new_entries_added = False
    for image_data in all_images:
        unique_id = image_data['id']
        
        # If the task is not in the Excel file, add it
        if unique_id not in existing_ids:
            new_entries_added = True
            image_path = image_data['path']
            qc_ref = image_data['qc_ref']
            image_name = os.path.basename(image_path)
            folder_path = os.path.dirname(image_path)
            folder_name = os.path.basename(folder_path)
            
            # Extract date from image name
            parts = image_name.split('_')
            if len(parts) > 2:
                relevant_name = parts[1]
                month, year = relevant_name[:3], relevant_name[3:]
                date_str = f"{month_to_number(month)}/{year}"
            else:
                date_str = "00/0000"
            
            # Append a new row with empty values for annotation
            ws.append([folder_name, date_str, None, image_name, None, None, qc_ref, unique_id, image_data['order']])
            
    if new_entries_added:
        print(f"{len(all_images) - len(existing_ids)} new tasks found. Saving to Excel file...")
        save_workbook()
        print("Excel file updated.")
    else:
        print("No new images to add. Excel file is up to date.")

# Run the population function right at the start
populate_excel_on_startup()
# --- End of pre-population logic ---


# Initialize control variables
current_image_idx = initial_image_idx # Use the restored or initial index
current_image_name = ""
original_pil_image = None # <- holds the pristine, undrawn PIL image
current_pil_image = None  # <- holds the PIL image that can be drawn on
current_tk_image = None   # <- reference to the current PhotoImage (avoids GC)
image_canvas_image_id = None  # <- id of the image item on the Canvas
resize_job = None  # <- id for the after() job for debouncing
zoom_level = 1.0  # <- current zoom level
view_rect = [0, 0, 1, 1] # <- [x0, y0, x1, y1] of the visible area (in relative coordinates)
pan_start_x = 0
pan_start_y = 0
pan_start_view_rect = None
last_draw_coords = None # <- Last coordinates for drawing
has_drawn_on_current_image = False # <- Flag to check if drawing has occurred
current_drawing_data = [] # <- List to store vector drawing strokes

# Function to extract and convert the date from the image name
def extract_date(name):
    match = re.search(r'_(\D{3}\d{4})_', name)  # Searches for a pattern like "_oct1995_"
    if match:
        month_year = match.group(1)
        month, year = month_year[:3], month_year[3:]
        return year + month_to_number(month)  # Returns YYYYMM
    return "000000"  # If no date is found, returns a default value

# Function to load the current image
def load_current_image():
    global current_image_idx, current_image_name
    if not all_images:
        return None, None, None, None
    
    image_data = all_images[current_image_idx]
    image_path = image_data['path']
    qc_ref = image_data['qc_ref']
    unique_id = image_data['id']
    current_image_name = os.path.basename(image_path)
    return cv2.imread(image_path), current_image_name, qc_ref, unique_id

# Function to check if an entry for the image exists and return its cropmark and saved status
def check_entry_exists(unique_id):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Find row by unique ID
        if len(row) > 7 and row[7] == unique_id:
            cropmark = row[2]
            saved_status = row[4]
            # Check if the drawing column exists and has content
            drawing_data = row[5] if len(row) > 5 and row[5] else None
            return True, cropmark, saved_status, drawing_data
    return False, None, None, None

# Function to redraw the saved strokes on the image
def redraw_saved_drawing(drawing_data):
    if not drawing_data or current_pil_image is None:
        return

    img_w, img_h = current_pil_image.size

    def normalize(data):
        out = []
        for stroke in data:
            pts = []
            for pt in stroke:
                try:
                    x, y = pt
                    xi = max(0, min(img_w - 1, int(round(float(x)))))
                    yi = max(0, min(img_h - 1, int(round(float(y)))))
                    pts.append((xi, yi))  # ensure tuple of ints
                except Exception:
                    continue
            if len(pts) > 1:
                out.append(pts)
        return out

    draw = ImageDraw.Draw(current_pil_image)
    for stroke in normalize(drawing_data):
        draw.line(stroke, fill="red", width=12)

# Function to fit and show the image on the canvas, maintaining aspect ratio
def fit_and_show_image():
    global current_pil_image, current_tk_image, image_canvas_image_id, view_rect
    if current_pil_image is None:
        return

    canvas_w = image_canvas.winfo_width()
    canvas_h = image_canvas.winfo_height()
    if canvas_w <= 1 or canvas_h <= 1:
        return

    img_w, img_h = current_pil_image.size
    
    # Coordinates of the area to be cropped from the original image
    x0 = int(view_rect[0] * img_w)
    y0 = int(view_rect[1] * img_h)
    x1 = int(view_rect[2] * img_w)
    y1 = int(view_rect[3] * img_h)

    # Crop the region of interest
    cropped_img = current_pil_image.crop((x0, y0, x1, y1))

    # Resize the cropped image to fit the canvas
    ratio = min(canvas_w / cropped_img.width, canvas_h / cropped_img.height)
    new_w = max(1, int(cropped_img.width * ratio))
    new_h = max(1, int(cropped_img.height * ratio))

    # Choose resampling algorithm: BICUBIC for upscaling (zoom in), LANCZOS for downscaling.
    if new_w > cropped_img.width or new_h > cropped_img.height:
        resampling_algorithm = Image.BICUBIC
    else:
        resampling_algorithm = Image.LANCZOS

    resized = cropped_img.resize((new_w, new_h), resampling_algorithm)
    current_tk_image = ImageTk.PhotoImage(resized)

    cx = canvas_w // 2
    cy = canvas_h // 2
    if image_canvas_image_id is None:
        image_canvas_image_id = image_canvas.create_image(cx, cy, image=current_tk_image, anchor="center")
    else:
        image_canvas.itemconfig(image_canvas_image_id, image=current_tk_image)
        image_canvas.coords(image_canvas_image_id, cx, cy)

# Zoom function with the mouse scroll
def zoom(event):
    global zoom_level, view_rect
    
    # Zoom factor
    zoom_factor = 1.1 if event.delta > 0 else 1 / 1.1
    
    # Limit zoom out to the original size
    if zoom_level * zoom_factor < 1:
        zoom_level = 1.0
        view_rect = [0, 0, 1, 1]
    else:
        zoom_level *= zoom_factor

        # Mouse coordinates on the canvas (0 to 1)
        mouse_x = event.x / image_canvas.winfo_width()
        mouse_y = event.y / image_canvas.winfo_height()

        # Current width and height of the view
        view_w = view_rect[2] - view_rect[0]
        view_h = view_rect[3] - view_rect[1]

        # Point on the original image corresponding to the mouse
        img_x = view_rect[0] + mouse_x * view_w
        img_y = view_rect[1] + mouse_y * view_h

        # New width and height of the view
        new_view_w = 1 / zoom_level
        new_view_h = 1 / zoom_level

        # New view coordinates, centered on the mouse point
        new_x0 = img_x - mouse_x * new_view_w
        new_y0 = img_y - mouse_y * new_view_h
        new_x1 = new_x0 + new_view_w
        new_y1 = new_y0 + new_view_h

        # Keep the view within the image boundaries (0 to 1)
        if new_x0 < 0: new_x0 = 0
        if new_y0 < 0: new_y0 = 0
        if new_x1 > 1: new_x1 = 1; new_x0 = new_x1 - new_view_w
        if new_y1 > 1: new_y1 = 1; new_y0 = new_y1 - new_view_h
        
        view_rect = [new_x0, new_y0, new_x1, new_y1]

    fit_and_show_image()

# Functions for drawing on the image
def canvas_to_image_coords(canvas_x, canvas_y):
    """Converts Canvas coordinates to original image coordinates."""
    if current_pil_image is None:
        return None

    img_w, img_h = current_pil_image.size
    canvas_w = image_canvas.winfo_width()
    canvas_h = image_canvas.winfo_height()

    # Coordinates of the visible area on the original image
    view_x0_img = view_rect[0] * img_w
    view_y0_img = view_rect[1] * img_h
    view_x1_img = view_rect[2] * img_w
    view_y1_img = view_rect[3] * img_h
    view_w_img = view_x1_img - view_x0_img
    view_h_img = view_y1_img - view_y0_img

    # Aspect ratio of the image displayed on the canvas
    ratio = min(canvas_w / view_w_img, canvas_h / view_h_img)
    displayed_w = view_w_img * ratio
    displayed_h = view_h_img * ratio

    # "Black bars" or offset of the centered image on the canvas
    offset_x = (canvas_w - displayed_w) / 2
    offset_y = (canvas_h - displayed_h) / 2

    # Click coordinates relative to the displayed image (not the canvas)
    click_x_on_displayed = canvas_x - offset_x
    click_y_on_displayed = canvas_y - offset_y

    # Convert to coordinates on the original image
    final_x = view_x0_img + (click_x_on_displayed / ratio)
    final_y = view_y0_img + (click_y_on_displayed / ratio)

    # Ensure coordinates are within the image boundaries
    if 0 <= final_x < img_w and 0 <= final_y < img_h:
        return (final_x, final_y)
    return None

def start_draw(event):
    """Starts the drawing process and a new stroke."""
    global last_draw_coords, current_drawing_data
    # Start a new stroke (a new list of points)
    current_drawing_data.append([])
    last_draw_coords = canvas_to_image_coords(event.x, event.y)
    if last_draw_coords:
        # Add the first point to the new stroke
        current_drawing_data[-1].append(last_draw_coords)

def draw_on_image(event):
    """Draws a line on the PIL image and saves the coordinates."""
    global last_draw_coords, has_drawn_on_current_image, current_drawing_data
    if last_draw_coords is None:
        return

    current_coords = canvas_to_image_coords(event.x, event.y)
    if current_coords:
        # Add the new point to the current stroke
        current_drawing_data[-1].append(current_coords)

        # Clear the prompt message and enable delete button when drawing starts
        if not has_drawn_on_current_image:
            draw_prompt_label.config(text="")
            delete_drawing_button.config(state="normal")

        draw = ImageDraw.Draw(current_pil_image)
        # Draws a red line with 12px thickness
        draw.line([last_draw_coords, current_coords], fill="red", width=12)
        last_draw_coords = current_coords
        
        # Mark that drawing has occurred and enable the save button
        has_drawn_on_current_image = True
        save_button.config(state="normal")

        # Update the image on the canvas to show the drawing
        fit_and_show_image()

def stop_draw(event):
    """Ends the drawing process."""
    global last_draw_coords
    last_draw_coords = None

# Functions to enable/disable drawing
def enable_drawing():
    image_canvas.bind("<ButtonPress-3>", start_draw)
    image_canvas.bind("<B3-Motion>", draw_on_image)
    image_canvas.bind("<ButtonRelease-3>", stop_draw)

def disable_drawing():
    image_canvas.unbind("<ButtonPress-3>")
    image_canvas.unbind("<B3-Motion>")
    image_canvas.unbind("<ButtonRelease-3>")

# Function to apply contrast/sharpness and redraw drawings
def apply_enhancements(value=None):
    global current_pil_image
    if original_pil_image is None:
        return
    
    contrast_level = contrast_slider.get()
    brightness_level = brightness_slider.get()
    
    # Start with the original image
    enhanced_image = original_pil_image
    
    # Apply Contrast
    enhancer_contrast = ImageEnhance.Contrast(enhanced_image)
    enhanced_image = enhancer_contrast.enhance(contrast_level)
    
    # Apply Brightness
    enhancer_brightness = ImageEnhance.Brightness(enhanced_image)
    enhanced_image = enhancer_brightness.enhance(brightness_level)
    
    # Set the enhanced image as the current one to be worked on
    current_pil_image = enhanced_image
    
    # Re-apply any existing drawings on top of the contrast-adjusted image
    if has_drawn_on_current_image:
        redraw_saved_drawing(current_drawing_data)
        
    # Update the canvas
    fit_and_show_image()

# Function to delete the current drawing
def delete_drawing():
    global current_pil_image, has_drawn_on_current_image, current_drawing_data
    if not has_drawn_on_current_image:
        return

    # First, clear the drawing data and reset the flag
    current_drawing_data = []
    has_drawn_on_current_image = False
    
    # Now, restore the image with current enhancement levels.
    # Since has_drawn_on_current_image is False, it will not redraw the old lines.
    apply_enhancements()

    # Update UI state
    delete_drawing_button.config(state="disabled")
    save_button.config(state="disabled")
    draw_prompt_label.config(text="Please draw the mark with the right mouse button.")

# Functions for panning (dragging the image)
def start_pan(event):
    global pan_start_x, pan_start_y, pan_start_view_rect
    pan_start_x = event.x
    pan_start_y = event.y
    pan_start_view_rect = list(view_rect) # Save a copy
    image_canvas.config(cursor="fleur")

def pan_image(event):
    global view_rect
    if pan_start_view_rect is None:
        return

    # Current width and height of the view in image coordinates (0 to 1)
    view_w = pan_start_view_rect[2] - pan_start_view_rect[0]
    view_h = pan_start_view_rect[3] - pan_start_view_rect[1]

    # Mouse displacement in pixels
    delta_x_pixels = event.x - pan_start_x
    delta_y_pixels = event.y - pan_start_y

    # Convert pixel displacement to image coordinates
    # (how much of the total image fits on the canvas)
    img_w, img_h = current_pil_image.size
    canvas_w = image_canvas.winfo_width()
    canvas_h = image_canvas.winfo_height()
    
    # Calculate the scale of the current view to the canvas
    scale = view_w * img_w / canvas_w
    
    delta_x = delta_x_pixels * scale / img_w
    delta_y = delta_y_pixels * scale / img_h

    # Calculate new view position
    new_x0 = pan_start_view_rect[0] - delta_x
    new_y0 = pan_start_view_rect[1] - delta_y
    new_x1 = new_x0 + view_w
    new_y1 = new_y0 + view_h

    # Keep the view within the image boundaries
    if new_x0 < 0: new_x0 = 0; new_x1 = view_w
    if new_y0 < 0: new_y0 = 0; new_y1 = view_h
    if new_x1 > 1: new_x1 = 1; new_x0 = 1 - view_w
    if new_y1 > 1: new_y1 = 1; new_y0 = 1 - view_h

    view_rect = [new_x0, new_y0, new_x1, new_y1]
    fit_and_show_image()

def end_pan(event):
    global pan_start_view_rect
    pan_start_view_rect = None
    image_canvas.config(cursor="")

# Debounce for canvas resizing
def on_canvas_resize(event):
    global resize_job
    if resize_job is not None:
        image_canvas.after_cancel(resize_job)
    resize_job = image_canvas.after(60, fit_and_show_image)

# Function to update the interface with the current image
def update_image():
    global original_pil_image, current_pil_image, zoom_level, view_rect, has_drawn_on_current_image, current_drawing_data
    image, image_name, qc_ref, unique_id = load_current_image()
    if image is None:
        print("No image found.")
        return

    # Reset zoom, drawing state, and vector data for the new image
    zoom_level = 1.0
    view_rect = [0, 0, 1, 1]
    has_drawn_on_current_image = False
    current_drawing_data = []
    draw_prompt_label.config(text="") # Clear the prompt message
    delete_drawing_button.config(state="disabled") # Disable delete button
    contrast_slider.set(1.5) # Reset contrast slider
    brightness_slider.set(1.2) # Reset brightness slider

    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    original_pil_image = Image.fromarray(image_rgb)
    # current_pil_image is now set inside apply_enhancements

    # Check if an entry exists and get the values using the unique ID
    entry_exists, cropmark_value, saved_status, drawing_json = check_entry_exists(unique_id)

    # If a saved drawing exists, load it into the state
    saved_drawing_data = None
    if drawing_json:
        try:
            if isinstance(drawing_json, str):
                drawing_json_str = drawing_json.strip()
                if drawing_json_str:
                    saved_drawing_data = json.loads(drawing_json_str)
            elif isinstance(drawing_json, (list, tuple)):
                saved_drawing_data = drawing_json
        except Exception as e:
            log(f"Failed to parse drawing JSON for {image_name}: {e}\n{traceback.format_exc()}")

    if saved_drawing_data:
        current_drawing_data = saved_drawing_data
        has_drawn_on_current_image = True
        delete_drawing_button.config(state="normal")  # Enable delete if drawing is loaded

    # Apply default enhancements which creates the initial current_pil_image
    # This function will now also handle redrawing the loaded data.
    apply_enhancements()

    # The rest of the UI updates can now proceed
    parts = image_name.split('_')
    if len(parts) > 2:
        relevant_name = parts[1]
    else:
        relevant_name = image_name

    # Update labels
    image_data = all_images[current_image_idx]
    folder_path = os.path.dirname(image_data['path'])
    folder_name = os.path.basename(folder_path)
    # image_name_label.config(text=f"Date: {relevant_name}", font=("Arial", 24))
    # folder_name_label.config(text=f"Site: {folder_name}", font=("Arial", 24), wraplength=400)

    # Force the user to choose an option
    save_button.config(state="disabled")
    mark_option_var.set(99) # Value that doesn't correspond to any option

    # Enable/disable drawing based on the existing annotation
    if entry_exists and cropmark_value is not None and int(cropmark_value) in [1, 2]:
        enable_drawing()
    else:
        disable_drawing()

    if entry_exists:
        if cropmark_value is not None:
            mark_option_var.set(int(cropmark_value))
            # If the loaded option is for drawing, check if a drawing exists
            if int(cropmark_value) in [1, 2] and has_drawn_on_current_image:
                 save_button.config(state="normal")
            elif int(cropmark_value) == 0:
                 save_button.config(state="normal")

        if saved_status == 'S':
            save_button.config(state="disabled")
    
    # Update the appearance of the radio buttons
    update_radio_buttons()

# Function to advance to the next image
def next_image():
    global current_image_idx
    if current_image_idx < len(all_images) - 1:
        current_image_idx += 1
    else:
        current_image_idx = 0 # Go back to the start if at the end
    update_image()

# Function to go back to the previous image
def prev_image():
    global current_image_idx
    if current_image_idx > 0:
        current_image_idx -= 1
    else:
        current_image_idx = len(all_images) - 1 # Go to the end if at the start
    update_image()

# Function to save the annotation
def save_annotation():
    global current_image_name
    image, image_name, qc_ref, unique_id = load_current_image() # Get all data
    if image is None:
        return
    cropmark = mark_option_var.get()
    image_data = all_images[current_image_idx]
    folder_path = os.path.dirname(image_data['path'])
    folder_name = os.path.basename(folder_path)
    parts = image_name.split('_')
    if len(parts) > 2:
        relevant_name = parts[1]
        month, year = relevant_name[:3], relevant_name[3:]
        date_str = f"{month_to_number(month)}/{year}"
    else:
        date_str = "00/0000"

    # Check if the entry already exists
    entry_exists = False
    drawing_json = json.dumps(current_drawing_data) if current_drawing_data else ""

    # Find the correct row using the unique ID and update it
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if len(row) > 7 and row[7].value == unique_id:
            ws.cell(row=row_idx, column=3, value=cropmark)
            ws.cell(row=row_idx, column=5, value='S')
            ws.cell(row=row_idx, column=6, value=drawing_json)
            # QC ref and unique ID are already set during pre-population, but we can ensure they are correct
            ws.cell(row=row_idx, column=7, value=qc_ref)
            entry_exists = True
            break

    if not entry_exists:
        # This case should not happen with the new pre-population logic
        print(f"Warning: Could not find pre-populated row for ID {unique_id}. Appending new row.")
        ws.append([folder_name, date_str, cropmark, image_name, 'S', drawing_json, qc_ref, unique_id])

    save_workbook()

    # Disable the "Save" button after clicking
    save_button.config(state="disabled")

    # Update the image counter
    update_counter()

    # Advance to the next image
    next_image()

# Function to enable the Save button and manage the 'S' state
def on_option_select():
    # This function now only handles removing the 'S' when changing options
    # Button activation is handled in on_radio_click and draw_on_image
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[3].value == current_image_name:
            if row[4].value == 'S':
                ws.cell(row=row_idx, column=5, value=None)
                save_workbook()
            break

# Function to update the appearance of custom radio buttons
def update_radio_buttons():
    current_value = mark_option_var.get()
    for i, label in enumerate(option_labels):
        if i == current_value:
            label.config(image=checked_image)
        else:
            label.config(image=unchecked_image)

# Function to handle clicks on a custom radio button
def on_radio_click(value):
    mark_option_var.set(value)
    update_radio_buttons()
    on_option_select()

    if value == 0: # "No Mark"
        save_button.config(state="normal")
        disable_drawing()
        draw_prompt_label.config(text="") # Clear the message
    elif value in [1, 2]: # "Faint Mark" or "Clear Mark"
        enable_drawing()
        # Disable the button and show the prompt until a drawing is made
        if not has_drawn_on_current_image:
            save_button.config(state="disabled")
            draw_prompt_label.config(text="Please draw the mark with the right mouse button.")
        else:
            save_button.config(state="normal")
            draw_prompt_label.config(text="")
    else:
        disable_drawing()
        draw_prompt_label.config(text="")

# Function to calculate the total number of images
def get_total_images_count():
    return len(all_images)

# Function to count saved images
def get_saved_images_count():
    saved_count = 0
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        if row[0] == 'S':
            saved_count += 1
    return saved_count

# Function to update the counter
def update_counter():
    total_images = get_total_images_count()
    saved_images = get_saved_images_count()
    remaining = total_images - saved_images
    counter_label.config(text=f"Images remaining: {remaining} / {total_images}")

# Function to create a custom dialog box
def custom_askokcancel(title, message):
    dialog = Toplevel(root)
    dialog.title(title)
    dialog.resizable(False, False)

    # Store the result
    dialog.result = False

    def on_ok():
        dialog.result = True
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    # Message
    msg_frame = Frame(dialog, padx=20, pady=20)
    msg_frame.pack()
    msg_label = Label(msg_frame, text=message, font=("Arial", 14), justify='left')
    msg_label.pack()

    # Buttons
    btn_frame = Frame(dialog, pady=10)
    btn_frame.pack()

    ok_button = Button(btn_frame, text="Exit", command=on_ok, font=("Arial", 12), width=10)
    ok_button.pack(side="left", padx=10)
    cancel_button = Button(btn_frame, text="Cancel", command=on_cancel, font=("Arial", 12), width=10)
    cancel_button.pack(side="right", padx=10)

    # Make the window modal and center it
    dialog.transient(root)
    dialog.update_idletasks()
    x = root.winfo_x() + (root.winfo_width() // 2) - (dialog.winfo_width() // 2)
    y = root.winfo_y() + (root.winfo_height() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f"+{x}+{y}")
    dialog.grab_set()
    
    root.wait_window(dialog)
    return dialog.result

# Function to close the application
def exit_app():
    # Ask for confirmation
    total_images = get_total_images_count()
    saved_images = get_saved_images_count()
    remaining = total_images - saved_images
    
    message = (
        "Are you sure you want to exit?\n\n"
        f"{remaining} images left to label.\n"
        "(Your progress is saved in the Excel file)"
    )
    
    if custom_askokcancel("Exit", message):
        root.destroy()

# Function to show the instructions dialog
def show_help_dialog():
    help_dialog = Toplevel(root)
    help_dialog.title("Instructions")
    help_dialog.resizable(False, False)

    try:
        # Load the instructions image from the 'resources' sub-directory
        instructions_img_path = os.path.join(script_dir, "resources", "instructions.png")
        instructions_img_pil = Image.open(instructions_img_path)
        
        # --- FIX: Apply conditional scaling based on the detected display scaling ---
        reduction_factor = 0.3 # Keep the 30% reduction.
        
        # Apply a different softening factor based on the display scaling
        if scaling_factor == 1.5:
            softened_scaling_factor = scaling_factor * 0.5
        elif scaling_factor == 1.25:
            softened_scaling_factor = scaling_factor * 0.7
        else:
            # Default behavior for other scaling levels (e.g., 1.0, 1.75)
            softened_scaling_factor = scaling_factor 
        
        new_width = int(instructions_img_pil.width / softened_scaling_factor * reduction_factor)
        new_height = int(instructions_img_pil.height / softened_scaling_factor * reduction_factor)
        
        # Use LANCZOS for high-quality downscaling
        instructions_img_pil = instructions_img_pil.resize((new_width, new_height), Image.LANCZOS)

        instructions_img_tk = ImageTk.PhotoImage(instructions_img_pil)
        
        # Display the image
        img_label = Label(help_dialog, image=instructions_img_tk)
        img_label.image = instructions_img_tk # Keep a reference to avoid garbage collection
        img_label.pack()

    except FileNotFoundError:
        # Fallback text if the image is not found, now using a scaled font
        Label(help_dialog, text=f"'instructions.png' not found in 'resources' directory.", font=font_tiny).pack(padx=40, pady=20)

    # Add a close button, now using a scaled font
    close_button = Button(help_dialog, text="Close", command=help_dialog.destroy, font=font_medium)
    close_button.pack(pady=10)

    # Make the window modal and center it
    help_dialog.transient(root)
    help_dialog.update_idletasks()
    x = root.winfo_x() + (root.winfo_width() // 2) - (help_dialog.winfo_width() // 2)
    y = root.winfo_y() + (root.winfo_height() // 2) - (help_dialog.winfo_height() // 2)
    help_dialog.geometry(f"+{x}+{y}")
    help_dialog.grab_set()
    root.wait_window(help_dialog)

# Create the graphical interface
root = Tk()
root.title("cropmarker v3.0")

# --- DPI Scaling Detection and Font Adjustment ---
# Get the scaling factor by comparing the screen's DPI to the standard 96 DPI
# This must be done after the root window is created.
try:
    # This is the most reliable way to get the scaling factor on Windows
    from ctypes import windll
    # FIX: This line can interfere with drawing libraries in a packaged EXE.
    # Commenting it out allows drawing to function correctly.
    # windll.shcore.SetProcessDpiAwareness(1) 
    scaling_factor = windll.shcore.GetScaleFactorForDevice(0) / 100
except (ImportError, AttributeError):
    # Fallback for non-Windows or if ctypes fails
    scaling_factor = root.winfo_fpixels('1i') / 96.0

print(f"Detected display scaling: {scaling_factor:.2f}")

# Define base font sizes (what looks good at 100% scaling)
BASE_FONT_SIZE_LARGE = 30
BASE_FONT_SIZE_MEDIUM = 24
BASE_FONT_SIZE_SMALL = 18
BASE_FONT_SIZE_TINY = 14

# Calculate scaled font sizes
font_large = ("Arial", int(BASE_FONT_SIZE_LARGE / scaling_factor))
font_medium = ("Arial", int(BASE_FONT_SIZE_MEDIUM / scaling_factor))
font_small = ("Arial", int(BASE_FONT_SIZE_SMALL / scaling_factor))
font_tiny = ("Arial", int(BASE_FONT_SIZE_TINY / scaling_factor))
# --- End of DPI Scaling ---


# --- Conditional Window Sizing ---
# Get screen dimensions to check for ultrawide monitors
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# If an ultrawide 3440x1440 screen is detected, use a fixed 16:9 window.
if screen_width == 3440 and screen_height == 1440:
    print("Ultrawide screen detected. Setting window to 1920x1080.")
    root.geometry("1920x1080")
    root.resizable(False, False)
else:
    # For all other resolutions, open maximized.
    root.resizable(True, True)
    root.state('zoomed')
# --- End of Conditional Sizing ---


# Handle the window close button ('X') to save the session
root.protocol("WM_DELETE_WINDOW", exit_app)

# Load images for custom radio buttons from the 'resources' sub-directory
unchecked_image_path = os.path.join(script_dir, "resources", "unchecked.png")
checked_image_path = os.path.join(script_dir, "resources", "checked.png")
unchecked_image = ImageTk.PhotoImage(Image.open(unchecked_image_path))
checked_image = ImageTk.PhotoImage(Image.open(checked_image_path))

# Create the IntVar for the Radiobuttons
mark_option_var = IntVar(value=99) # Starts with a value that doesn't match any option

# --- NEW, UNIFIED GRID LAYOUT ---

# Configure the root window's grid system
root.grid_rowconfigure(0, weight=1)    # Main content row (image/controls) will expand
root.grid_rowconfigure(1, weight=0)    # Bottom slider row will NOT expand
root.grid_columnconfigure(0, weight=1) # Image column will expand
root.grid_columnconfigure(1, weight=0) # Control column will NOT expand

# Image frame (placed in the grid)
image_frame = Frame(root)
image_frame.grid(row=0, column=0, sticky="nsew")

# Controls frame (placed in the grid)
control_frame = Frame(root)
control_frame.grid(row=0, column=1, sticky="n", padx=35) # Aligns to the top

# Bottom frame for sliders (placed in the grid, spanning both columns)
bottom_frame = Frame(root)
bottom_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=20, pady=10)

# --- END OF LAYOUT DEFINITION ---


# Replace Label with Canvas for image drawing
image_canvas = Canvas(image_frame, highlightthickness=0)
image_canvas.pack(fill="both", expand=True) # Canvas fills its parent frame

# Resize the image whenever the Canvas size changes (with debounce)
image_canvas.bind("<Configure>", on_canvas_resize)
image_canvas.bind("<MouseWheel>", zoom)
# Panning with left button
image_canvas.bind("<ButtonPress-1>", start_pan)
image_canvas.bind("<B1-Motion>", pan_image)
image_canvas.bind("<ButtonRelease-1>", end_pan)
# Drawing with right button (now enabled/disabled dynamically)
# image_canvas.bind("<ButtonPress-3>", start_draw)
# image_canvas.bind("<B3-Motion>", draw_on_image)
# image_canvas.bind("<ButtonRelease-3>", stop_draw)

# NOTE: The control_frame is now created and placed by the main grid.

# Frame for navigation buttons
navigation_frame = Frame(control_frame)
navigation_frame.pack(pady=50)

# Side-by-side navigation buttons
prev_button = Button(navigation_frame, text="Previous (A)", command=prev_image, font=font_medium, width=10, height=2)
prev_button.pack(side="left", padx=15)
next_button = Button(navigation_frame, text="Next (D)", command=next_image, font=font_medium, width=10, height=2)
next_button.pack(side="left", padx=15)

# Frame for mark options
options_frame = Frame(control_frame)
options_frame.pack(pady=1, anchor="w")

# Create custom radio buttons
option_labels = []
options = [("No Mark", 0), ("Faint Mark", 1), ("Clear Mark", 2)]

for text, value in options:
    frame = Frame(options_frame, bg=control_frame.cget('bg'))
    frame.pack(anchor="w", padx=30, pady=10)
    
    img_label = Label(frame, image=unchecked_image, bg=control_frame.cget('bg'))
    img_label.pack(side="left")
    
    text_label = Label(frame, text=text, font=font_large, bg=control_frame.cget('bg'))
    text_label.pack(side="left", padx=25)
    
    # Use a lambda function to capture the correct value
    img_label.bind("<Button-1>", lambda e, v=value: on_radio_click(v))
    text_label.bind("<Button-1>", lambda e, v=value: on_radio_click(v))
    
    option_labels.append(img_label)

# Prompt message for drawing
draw_prompt_label = Label(control_frame, text="", font=font_tiny, fg="red")
draw_prompt_label.pack(pady=10)

# Delete Drawing Button
delete_drawing_button = Button(control_frame, text="Delete Drawing (Del)", command=delete_drawing, font=font_small, state="disabled")
delete_drawing_button.pack(pady=10)

# Save button
save_button = Button(control_frame, text="Save (S)", command=save_annotation, font=font_medium, width=10, height=2)
save_button.pack(pady=20)

# Label for the counter
counter_label = Label(control_frame, text="", font=font_large)
counter_label.pack(pady=20)

# Exit button
exit_button = Button(control_frame, text="Exit (Esc)", command=exit_app, font=font_medium, width=10, height=2)
exit_button.pack(side="bottom", pady=20)

# NOTE: The bottom_frame is now created and placed by the main grid.

# Help Button ('?') in the bottom-right corner
help_button = Button(bottom_frame, text="?", font=("Arial", int(BASE_FONT_SIZE_MEDIUM / scaling_factor), "bold"), command=show_help_dialog)
help_button.pack(side="right", padx=20)

# A central frame for the sliders, which will be centered
sliders_container = Frame(bottom_frame)
sliders_container.pack(expand=True) # Center this container

# Contrast Slider Frame
contrast_frame = Frame(sliders_container)
contrast_frame.pack(side="left", padx=20) # Pack side-by-side
Label(contrast_frame, text="Contrast", font=font_medium).pack(side="left", padx=(0, 10))
contrast_slider = Scale(contrast_frame, from_=0.5, to=3.0, resolution=0.1, orient="horizontal", command=apply_enhancements, length=400)
contrast_slider.set(1.5)
contrast_slider.pack(side="left") # Don't expand or fill

# Brightness Slider Frame
brightness_frame = Frame(sliders_container)
brightness_frame.pack(side="left", padx=20) # Pack side-by-side
Label(brightness_frame, text="Brightness", font=font_medium).pack(side="left", padx=(0, 10))
brightness_slider = Scale(brightness_frame, from_=0.5, to=2.0, resolution=0.1, orient="horizontal", command=apply_enhancements, length=400)
brightness_slider.set(1.2)
brightness_slider.pack(side="left") # Don't expand or fill

# Function to handle initial setup after the main window is ready
def initial_setup():
    # Load the first image and initialize the counter
    update_image()
    update_counter()
    # Now that the app is ready, show the help dialog
    show_help_dialog()

# Add keyboard shortcuts
root.bind('<a>', lambda event: prev_image())
root.bind('<d>', lambda event: next_image())
root.bind('<s>', lambda event: save_annotation() if save_button['state'] == 'normal' else None)
root.bind('<Delete>', lambda event: delete_drawing())
root.bind('<Escape>', lambda event: exit_app())

# Schedule the initial setup to run once the main loop is idle
root.after(100, initial_setup)

# Start the main loop
root.mainloop()