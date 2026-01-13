import os
import cv2
import openpyxl
from tkinter import Tk, Label, Button, Frame, Canvas, StringVar, OptionMenu, Toplevel, Scrollbar, LabelFrame, BooleanVar, Checkbutton, DoubleVar, Scale
from tkinter.ttk import Combobox
from PIL import Image, ImageTk, ImageDraw
import re
import json
import statistics
import glob
import sys
import csv
import pandas as pd
import numpy as np
from collections import defaultdict

# --- Configuration ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))

class Config:
    """Holds all static configuration variables for the application."""
    MAIN_FOLDER = os.path.join(SCRIPT_DIR, 'image_dataset')
    ANNOTATIONS_FOLDER = os.path.join(SCRIPT_DIR, 'annotations')
    VALIDATIONS_FILE = os.path.join(ANNOTATIONS_FOLDER, 'supervisor_validations.json')
    DS_LOG_FILE = os.path.join(ANNOTATIONS_FOLDER, 'ds_convergence_log.txt')
    RATERS_REPORT_FILE = os.path.join(ANNOTATIONS_FOLDER, 'raters_report.csv')
    CONSENSUS_REPORT_FILE = os.path.join(ANNOTATIONS_FOLDER, 'consensus.csv')

    USER_COLORS = ['#FF0000', '#00FF00', '#0000FF', '#FFFF00', '#FF00FF', '#00FFFF', '#FFA500', '#800080']
    
    # Dawid-Skene model thresholds
    DS_PRESENCE_THRESHOLD = 0.6
    DS_VALIDATION_LOWER_BOUND = 0.15
    DS_VALIDATION_UPPER_BOUND = 0.85

    # Supervisor validation weight parameters
    WEIGHT_BASELINE = 1.0
    WEIGHT_MIN_CLAMP = 0.1
    WEIGHT_MAX_CLAMP = 1.5
    
    VALIDATION_DELTAS = {
        "D1 - Growing vegetation vs bare soil": 0.2, "D2 - Healthier vegetation contrast": 0.2,
        "D3 - Dry vegetation with residual contrast": 0.2, "D4 - Faint micro-moisture variation": 0.2,
        "D5 - Aligned with known ditch plan": 0.4,
        "F1 - Machinery tracks": -0.3, "F2 - Boundaries, streams and drainage": -0.3,
        "F3 - Soil texture": -0.3, "F4 - Misalignment with known ditch plan": -0.3,
        "F5 - Fairy rings": -0.5,
        "M1 - Faint Mark": -0.2, "M2 - Clear Mark": -0.3
    }

    VALIDATION_FACTORS = {
        "Drivers": [
            "D1 - Growing vegetation vs bare soil", "D2 - Healthier vegetation contrast", 
            "D3 - Dry vegetation with residual contrast", "D4 - Faint micro-moisture variation",
            "D5 - Aligned with known ditch plan"
        ],
        "Absence Reason": ["I1 - Bare soil", "I2 - Dense uniform cover", "I3 - Low quality imagery"],
        "False Positives": [
            "F1 - Machinery tracks", "F2 - Boundaries, streams and drainage",
            "F3 - Soil texture", "F4 - Misalignment with known ditch plan",
            "F5 - Fairy rings"
        ],
        "Missed Positives": ["M1 - Faint Mark", "M2 - Clear Mark"]
    }

# --- Helper Classes ---
class Tooltip:
    """Create a tooltip for a given widget."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")

        label = Label(self.tooltip_window, text=self.text, justify='left',
                      background="#ffffe0", relief='solid', borderwidth=1,
                      font=("tahoma", "10", "normal"))
        label.pack(ipadx=1)

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

# --- Data Loading and Processing ---
def get_month_map():
    """Returns a dictionary mapping month abbreviations to numbers."""
    return {
        'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04', 'mai': '05', 'jun': '06',
        'jul': '07', 'ago': '08', 'set': '09', 'out': '10', 'nov': '11', 'dez': '12',
        'apr': '04', 'aug': '08', 'dec': '12', 'feb': '02', 'jan': '01', 'jul': '07',
        'jun': '06', 'mar': '03', 'may': '05', 'nov': '11', 'oct': '10', 'sep': '09'
    }

def extract_sort_key(image_path):
    """Extracts a sort key (folder, year, month) from an image path."""
    month_map = get_month_map()
    folder_name = os.path.basename(os.path.dirname(image_path))
    image_name = os.path.basename(image_path)
    
    match = re.search(r'_(\D{3})(\d{4})_', image_name)
    if match:
        month_str, year_str = match.groups()
        month_num = month_map.get(month_str.lower(), '00')
        return (folder_name, year_str, month_num)
    return (folder_name, "0000", "00")

def aggregate_user_data():
    """Finds all user excel files and aggregates their annotations and expertise scores."""
    user_files = glob.glob(os.path.join(Config.ANNOTATIONS_FOLDER, 'cropmarks_*.xlsx'))
    aggregated_data = {}
    user_color_map = {}
    users_with_unclassified = set()
    user_expertise_scores = {}

    print(f"Found {len(user_files)} user data files in '{Config.ANNOTATIONS_FOLDER}'.")

    for i, file_path in enumerate(user_files):
        filename = os.path.basename(file_path)
        # Accept usernames with underscores/spaces and Windows duplicate suffixes like " (1).xlsx".
        match = re.match(r'^cropmarks_(.+?)(?:_(\d+))?(?:\s*\(\d+\))?\.xlsx$', filename)
        if not match:
            print(f"Skipping unrecognized file name format: {filename}")
            continue

        username, score_str = match.groups()
        score = int(score_str) if score_str else 1

        if username not in user_color_map:
            user_color_map[username] = Config.USER_COLORS[len(user_color_map) % len(Config.USER_COLORS)]
            user_expertise_scores[username] = score

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            if any(row[4] != 'S' for row in ws.iter_rows(min_row=2, values_only=True) if len(row) > 4):
                users_with_unclassified.add(username)

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Expected columns (export-xlsx):
                # 0 Site, 1 Date, 2 Cropmark, 3 Image, 4 Saved, 5 Drawing, 6 QC_Reference, 7 UniqueID, 8 Order
                if len(row) > 4 and row[4] == 'S':
                    image_filename = row[3] if len(row) > 3 else None
                    instance_id = row[7] if len(row) > 7 else None
                    if not image_filename:
                        # Fall back to instance_id if needed.
                        image_filename = instance_id
                    if not image_filename:
                        continue

                    if image_filename not in aggregated_data:
                        aggregated_data[image_filename] = []

                    aggregated_data[image_filename].append({
                        'user': username,
                        'score': int(row[2]) if row[2] is not None else -1,
                        'image': row[3],
                        'drawing': row[5],
                        'qc_ref': row[6],
                        # Keep the exported UniqueID available for QC/traceability.
                        'instance_id': instance_id,
                    })
        except Exception as e:
            print(f"Could not process file {file_path}: {e}")
            
    return aggregated_data, user_color_map, users_with_unclassified, user_expertise_scores

# --- Dawid-Skene Implementation ---
class DawidSkene:
    """Implementation of the Dawid-Skene model using Expectation-Maximization."""
    def __init__(self, data, user_weights, user_consistency, observation_weights, log_file_path=None):
        self.data = data
        self.user_weights = user_weights
        self.user_consistency = user_consistency
        self.observation_weights = observation_weights
        self.log_file_path = log_file_path
        
        self.items = sorted(list(set(item for item, _, _ in self.data)))
        self.users = sorted(list(set(user for _, user, _ in self.data)))
        self.labels = sorted(list(set(label for _, _, label in self.data)))
        
        self.item_map = {item: i for i, item in enumerate(self.items)}
        self.user_map = {user: i for i, user in enumerate(self.users)}
        self.label_map = {label: j for j, label in enumerate(self.labels)}
        
        self.n_items = len(self.items)
        self.n_users = len(self.users)
        self.n_labels = len(self.labels)

        self._initialize_params()

    def _log(self, message):
        if self.log_file_path:
            with open(self.log_file_path, 'a') as f:
                f.write(message + '\n')

    def _initialize_params(self):
        """Initializes class prevalence and user confusion matrices with priors."""
        self.p = np.full(self.n_labels, 1.0 / self.n_labels)
        self.pi = np.zeros((self.n_users, self.n_labels, self.n_labels))
        for user, user_idx in self.user_map.items():
            weight = self.user_weights.get(user, 1.0)
            consistency = self.user_consistency.get(user, 1.0)
            prior_strength = (weight - 1.0) * 10 * consistency
            
            prior_matrix = np.full((self.n_labels, self.n_labels), 0.1)
            np.fill_diagonal(prior_matrix, 1.0)
            prior_matrix *= prior_strength
            prior_matrix += 0.5

            self.pi[user_idx] = prior_matrix / prior_matrix.sum(axis=1, keepdims=True)

    def run(self, max_iter=100, tol=1e-4):
        """Runs the EM algorithm until convergence."""
        if self.log_file_path and os.path.exists(self.log_file_path):
            os.remove(self.log_file_path)
        self._log("--- Dawid-Skene EM Convergence Log ---")

        last_log_likelihood = -np.inf
        for i in range(max_iter):
            T = self._e_step()
            self._m_step(T)
            
            log_likelihood = self._calculate_log_likelihood(T)
            self._log(f"Iteration {i+1}: Log-Likelihood = {log_likelihood:.6f}")
            if abs(log_likelihood - last_log_likelihood) < tol:
                self._log(f"\nEM converged after {i+1} iterations.")
                break
            last_log_likelihood = log_likelihood
        else:
            self._log(f"\nEM reached max iterations ({max_iter}).")
        
        self.posterior = self._e_step()
        return self.get_results()

    def _e_step(self):
        """Calculates the posterior probability T_ij that item i has true label j."""
        T = np.zeros((self.n_items, self.n_labels))
        for i, item in enumerate(self.items):
            for j, label in enumerate(self.labels):
                prod = 1.0
                for _, user, user_label in filter(lambda x: x[0] == item, self.data):
                    user_idx = self.user_map[user]
                    label_idx = self.label_map[user_label]
                    weight = self.observation_weights.get((item, user), 1.0)
                    prod *= (self.pi[user_idx, j, label_idx] ** weight)
                T[i, j] = self.p[j] * prod
        
        row_sums = T.sum(axis=1, keepdims=True)
        return np.divide(T, row_sums, out=np.full_like(T, 1/self.n_labels), where=row_sums!=0)

    def _m_step(self, T):
        """Updates model parameters based on the posterior probabilities T."""
        self.p = T.sum(axis=0) / self.n_items
        self.pi = np.zeros((self.n_users, self.n_labels, self.n_labels))
        for k, user in enumerate(self.users):
            user_annotations = [obs for obs in self.data if obs[1] == user]
            for j in range(self.n_labels):
                denominator = 0.0
                for item_id, _, _ in user_annotations:
                    item_idx = self.item_map[item_id]
                    weight = self.observation_weights.get((item_id, user), 1.0)
                    denominator += T[item_idx, j] * weight
                
                if denominator > 1e-9:
                    for l in range(self.n_labels):
                        numerator = 0.0
                        for item_id, _, obs_label in user_annotations:
                            if obs_label == self.labels[l]:
                                item_idx = self.item_map[item_id]
                                weight = self.observation_weights.get((item_id, user), 1.0)
                                numerator += T[item_idx, j] * weight
                        self.pi[k, j, l] = numerator / denominator
        
        for k in range(self.n_users):
            row_sums = self.pi[k].sum(axis=1, keepdims=True)
            self.pi[k] = np.divide(self.pi[k], row_sums, out=np.full_like(self.pi[k], 1/self.n_labels), where=row_sums!=0)

    def _calculate_log_likelihood(self, T):
        """Calculates the log-likelihood of the observed data given current parameters."""
        return np.sum(np.log(T.sum(axis=1)))

    def get_results(self):
        """Returns the final posterior probabilities for each item."""
        return {
            item: {self.labels[j]: self.posterior[i, j] for j in range(self.n_labels)}
            for i, item in enumerate(self.items)
        }

    def get_confusion_matrices(self):
        """Returns the final confusion matrices for all users."""
        return {user: self.pi[user_idx] for user, user_idx in self.user_map.items()}

    def compute_presence(self, weighted_median_map=None):
        """Returns per-item dict with ternary posteriors and binary presence."""
        out = {}
        if not np.array_equal(self.labels, [0, 1, 2]):
            print("Warning: DS labels are not [0, 1, 2]. Presence calculation may be incorrect.")
            return {}

        for i, item in enumerate(self.items):
            p0, p1, p2 = self.posterior[i]
            p_present = float(p1 + p2)
            present = (p_present >= Config.DS_PRESENCE_THRESHOLD)
            
            if weighted_median_map and weighted_median_map.get(item, 0) >= 1:
                present = True
            
            out[item] = {
                "P_none": float(p0), "P_faint": float(p1), "P_clear": float(p2),
                "P_present": p_present, "Present": bool(present)
            }
        return out

# --- End of Dawid-Skene Implementation ---
def calculate_weighted_median(annotations, user_effective_weights):
    """Calculates the weighted median for a single image's annotations."""
    if not annotations: return None
    weighted_scores = [(ann['score'], user_effective_weights.get(ann['user'], 1.0)) for ann in annotations if ann['score'] in [0, 1, 2]]
    if not weighted_scores: return None

    weighted_scores.sort(key=lambda x: x[0])
    total_weight = sum(w for _, w in weighted_scores)
    midpoint = total_weight / 2.0
    
    cumulative_weight = 0.0
    for score, weight in weighted_scores:
        cumulative_weight += weight
        if cumulative_weight >= midpoint:
            return score
    return weighted_scores[-1][0]

def calculate_user_consistency(aggregated_data):
    """Calculates user consistency based on consistency across QC duplicate images."""
    # We do NOT rely on the dict keys being QC IDs, because newer exports may
    # key by image filename. Instead, we use the per-row fields:
    # - instance_id (UniqueID column)
    # - qc_ref (QC_Reference column)
    qc_pairs = set()
    user_scores_by_instance = defaultdict(dict)

    for _, annotations in aggregated_data.items():
        for ann in annotations:
            inst = ann.get('instance_id') or ann.get('image')
            if inst:
                user_scores_by_instance[ann['user']][inst] = ann.get('score', -1)
            qc_ref = ann.get('qc_ref')
            if qc_ref and inst and qc_ref != inst:
                qc_pairs.add((qc_ref, inst))

    if not qc_pairs:
        print("No QC image pairs found to calculate consistency.")
        return {}

    user_consistency_scores = {}
    for user, scores in user_scores_by_instance.items():
        matches, total_annotated_pairs = 0, 0
        for original_id, duplicate_id in qc_pairs:
            if original_id in scores and duplicate_id in scores:
                total_annotated_pairs += 1
                score1, score2 = scores[original_id], scores[duplicate_id]
                if (score1 == 0 and score2 == 0) or (score1 > 0 and score2 > 0):
                    matches += 1
        
        consistency = (matches / total_annotated_pairs) if total_annotated_pairs > 0 else 1.0
        user_consistency_scores[user] = consistency
        print(f"User '{user}': consistency = {consistency:.2f} ({matches}/{total_annotated_pairs} QC pairs matched)")
    return user_consistency_scores

# --- Main Application Class ---
class SupervisorApp:
    """Main application class for the Supervisor GUI."""

    # --------------------------------------------------------------------------
    # 1. Initialization and Setup
    # --------------------------------------------------------------------------
    def __init__(self, root):
        self.root = root
        self.root.title("Supervisor GUI")
        self.root.geometry("1920x1080")
        self.root.resizable(False, False)

        self.all_image_paths, self.folder_list = self.get_image_and_folder_list()
        if not self.all_image_paths:
            print("No images found. Exiting."); self.root.destroy(); return
            
        self.aggregated_data, self.user_color_map, _, self.user_expertise_scores = aggregate_user_data()
        self.user_weights = {user: 1 + (score / 10.0) for user, score in self.user_expertise_scores.items()}
        self.user_consistency_scores = calculate_user_consistency(self.aggregated_data)
        self.supervisor_validations = self._load_validations()

        self.ds_results, self.ds_model = {}, None
        self._run_ds_model()

        # --- UI and State Variables ---
        self.global_image_idx = 0
        self.user_visibility = {user: True for user in self.user_color_map.keys()}
        self.show_heatmap = False
        self.zoom_level, self.view_rect = 1.0, [0, 0, 1, 1]
        self.pan_start_x, self.pan_start_y, self.pan_start_view_rect = 0, 0, None
        self.full_res_left_img, self.full_res_right_img = None, None
        
        self.validation_vars, self.submit_buttons = {}, {}
        self.validation_users_per_page, self.validation_current_page = 4, 0
        self.force_validation_override = False
        
        self.selected_folder = StringVar()
        self.folder_dropdown_programmatic_update = False
        
        self.setup_gui()
        self.root.bind('<a>', lambda e: self.prev_image())
        self.root.bind('<d>', lambda e: self.next_image())
        self.root.bind('<space>', lambda e: self.toggle_heatmap())
        
        self.update_display()

    def setup_gui(self):
        """Initializes all the main GUI widgets."""
        # Top Info and Navigation Frame
        self.info_frame = Frame(self.root, pady=10); self.info_frame.pack(fill='x', padx=20)
        self.site_label = Label(self.info_frame, text="Site: -", font=("Arial", 16)); self.site_label.pack(side='left')
        self.date_label = Label(self.info_frame, text="Date: -", font=("Arial", 16)); self.date_label.pack(side='left', padx=20)
        Frame(self.info_frame).pack(side='left', expand=True, fill='x') # Spacer

        self.prev_button = Button(self.info_frame, text="Previous (A)", command=self.prev_image, font=("Arial", 14), width=15); self.prev_button.pack(side='left', padx=10)
        
        folder_options = ["All Sites"] + self.folder_list
        self.folder_dropdown = Combobox(self.info_frame, textvariable=self.selected_folder, values=folder_options, state='readonly', font=("Arial", 12), width=20)
        self.selected_folder.set(folder_options[0])
        self.folder_dropdown.pack(side='left')
        self.folder_dropdown.bind("<MouseWheel>", self.on_scroll_folders)
        self.folder_dropdown.bind("<<ComboboxSelected>>", self.on_folder_select)

        self.next_button = Button(self.info_frame, text="Next (D)", command=self.next_image, font=("Arial", 14), width=15); self.next_button.pack(side='left', padx=10)
        self.heatmap_button = Button(self.info_frame, text="Toggle Heatmap (Space)", command=self.toggle_heatmap, font=("Arial", 14)); self.heatmap_button.pack(side='left', padx=20)

        # Image Canvases Frame
        self.image_frame = Frame(self.root); self.image_frame.pack(fill='both', expand=True, padx=20, pady=10)
        self.left_canvas = Canvas(self.image_frame, bg='black'); self.left_canvas.pack(side='left', fill='both', expand=True, padx=(0, 5))
        self.right_canvas = Canvas(self.image_frame, bg='black'); self.right_canvas.pack(side='right', fill='both', expand=True, padx=(5, 0))
        for canvas in [self.left_canvas, self.right_canvas]:
            canvas.bind("<ButtonPress-1>", self.start_pan); canvas.bind("<B1-Motion>", self.pan_image)
            canvas.bind("<ButtonRelease-1>", self.end_pan); canvas.bind("<MouseWheel>", self.zoom)

        # Bottom Frame for DS Results and Validation
        self.bottom_frame = Frame(self.root, pady=10); self.bottom_frame.pack(fill='x')
        ds_frame = Frame(self.bottom_frame); ds_frame.pack(fill='x', padx=20, pady=(0, 10))
        self.ds_results_label = Label(ds_frame, text="Certainty of Presence: -", font=("Arial", 14, "bold")); self.ds_results_label.pack(side='left')

        # Supervisor Validation Section
        self.validation_outer_frame = Frame(self.bottom_frame); self.validation_outer_frame.pack(side='top', anchor='nw', fill='both', expand=True, padx=20)
        validation_title_frame = Frame(self.validation_outer_frame); validation_title_frame.pack(fill='x', pady=(0, 5))
        Label(validation_title_frame, text="Supervisor Validation", font=("Arial", 12, "bold")).pack(side='left', anchor='w')

        pagination_controls_frame = Frame(validation_title_frame); pagination_controls_frame.pack(side='right')
        self.prev_page_button = Button(pagination_controls_frame, text="< Prev", command=self.prev_validation_page, font=("Arial", 9)); self.prev_page_button.pack(side='left', padx=5)
        self.page_label = Label(pagination_controls_frame, text="Page 1 of 1", font=("Arial", 9)); self.page_label.pack(side='left')
        self.next_page_button = Button(pagination_controls_frame, text="Next >", command=self.next_validation_page, font=("Arial", 9)); self.next_page_button.pack(side='left', padx=5)

        validation_canvas = Canvas(self.validation_outer_frame, highlightthickness=0)
        validation_scrollbar = Scrollbar(self.validation_outer_frame, orient="vertical", command=validation_canvas.yview)
        self.validation_scrollable_frame = Frame(validation_canvas)
        self.validation_scrollable_frame.bind("<Configure>", lambda e: validation_canvas.configure(scrollregion=validation_canvas.bbox("all")))
        validation_canvas.create_window((0, 0), window=self.validation_scrollable_frame, anchor="nw")
        validation_canvas.configure(yscrollcommand=validation_scrollbar.set)
        validation_canvas.pack(side="left", fill="both", expand=True)
        validation_scrollbar.pack(side="right", fill="y")

    # --------------------------------------------------------------------------
    # 2. Data Loading & Core Model Logic
    # --------------------------------------------------------------------------
    def get_image_and_folder_list(self):
        """Gets a sorted list of all .jpg images and a unique list of folder names."""
        folders = sorted([f for f in os.listdir(Config.MAIN_FOLDER) if os.path.isdir(os.path.join(Config.MAIN_FOLDER, f))])
        all_paths = [os.path.join(Config.MAIN_FOLDER, folder, img) for folder in folders for img in os.listdir(os.path.join(Config.MAIN_FOLDER, folder)) if img.lower().endswith('.jpg')]
        all_paths.sort(key=extract_sort_key)
        print(f"Found and sorted {len(all_paths)} unique images in {len(folders)} folders.")
        return all_paths, folders

    def _load_validations(self):
        """Loads supervisor validations from the JSON file."""
        try:
            if os.path.exists(Config.VALIDATIONS_FILE):
                with open(Config.VALIDATIONS_FILE, 'r') as f:
                    return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: Could not load or parse {os.path.basename(Config.VALIDATIONS_FILE)}: {e}")
        return {}

    def _run_ds_model(self):
        """Prepares data, runs the Dawid-Skene model, and generates reports."""
        consensus_data = defaultdict(list)
        for unique_id, annotations in self.aggregated_data.items():
            consensus_data[unique_id.replace('_qc', '')].extend(annotations)

        ds_data = [(item, ann['user'], ann['score']) for item, annotations in consensus_data.items() for ann in annotations if ann['score'] in [0, 1, 2]]
        if not ds_data:
            print("No data available to run Dawid-Skene model."); return

        observation_weights = {(img_id, user): val.get('weight', 1.0) for img_id, users in self.supervisor_validations.items() for user, val in users.items()}

        print("\nRunning Dawid-Skene model with supervisor weights...")
        self.ds_model = DawidSkene(ds_data, self.user_weights, self.user_consistency_scores, observation_weights, log_file_path=Config.DS_LOG_FILE)
        self.ds_model.run()
        
        user_effective_weights = {u: self.user_weights.get(u, 1.0) * self.user_consistency_scores.get(u, 1.0) for u in self.user_color_map}
        weighted_median_map = {item_id: calculate_weighted_median(ann, user_effective_weights) for item_id, ann in consensus_data.items()}
        self.ds_results = self.ds_model.compute_presence(weighted_median_map=weighted_median_map)
        
        print("Dawid-Skene model and presence calculation finished.")
        self._generate_reports(weighted_median_map)

    def _generate_reports(self, weighted_median_map):
        """Generates consensus and rater reports after the DS model runs."""
        # Raters Report
        confusion_matrices = self.ds_model.get_confusion_matrices()
        with open(Config.RATERS_REPORT_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['User', 'Expertise_Score', 'Calculated_consistency', 'Effective_Weight', 'T0_S0', 'T0_S1', 'T0_S2', 'T1_S0', 'T1_S1', 'T1_S2', 'T2_S0', 'T2_S1', 'T2_S2'])
            for user, matrix in sorted(confusion_matrices.items()):
                expertise = self.user_expertise_scores.get(user, 1)
                consistency = self.user_consistency_scores.get(user, 1.0)
                effective_weight = self.user_weights.get(user, 1.0) * consistency
                writer.writerow([user, expertise, f"{consistency:.4f}", f"{effective_weight:.4f}"] + list(matrix.flatten()))
        print(f"Raters report saved to {os.path.basename(Config.RATERS_REPORT_FILE)}")

        # Consensus Report
        with open(Config.CONSENSUS_REPORT_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Image_ID', 'P_Present', 'P_Faint', 'P_Clear', 'DS_Present', 'Weighted_Median_Label'])
            for image_id, results in sorted(self.ds_results.items()):
                writer.writerow([image_id, f"{results['P_present']:.4f}", f"{results['P_faint']:.4f}", f"{results['P_clear']:.4f}", results['Present'], weighted_median_map.get(image_id)])
        print(f"Consensus report saved to {os.path.basename(Config.CONSENSUS_REPORT_FILE)}")

    # --------------------------------------------------------------------------
    # 3. Main Display Logic
    # --------------------------------------------------------------------------
    def update_display(self, reset_view=True):
        """Loads the current image and its annotations and updates the entire GUI."""
        if not self.all_image_paths: return

        image_path = self.all_image_paths[self.global_image_idx]
        image_name = os.path.basename(image_path)
        
        if reset_view:
            self.zoom_level, self.view_rect = 1.0, [0, 0, 1, 1]
            self.validation_current_page, self.force_validation_override = 0, False
            try:
                cv_img = cv2.imread(image_path)
                if cv_img is None: raise IOError(f"Could not load image: {image_path}")
                self.full_res_left_img = Image.fromarray(cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB))
            except Exception as e:
                print(f"Error loading image: {e}"); self.full_res_left_img = None; return

        annotations = self.aggregated_data.get(image_name, [])
        original_image_name = image_name.replace('_qc', '')
        image_results = self.ds_results.get(original_image_name, {})
        p_present = image_results.get('P_present', 0.5)
        
        self._update_info_labels(image_path, p_present, image_results)
        self._update_validation_ui(annotations, original_image_name, p_present)
        self._update_right_canvas(annotations)
        
        self.fit_and_show_image(self.left_canvas, self.full_res_left_img)
        self.fit_and_show_image(self.right_canvas, self.full_res_right_img)

    def _update_info_labels(self, image_path, p_present, image_results):
        """Updates the top info labels (Site, Date, DS Score)."""
        site = os.path.basename(os.path.dirname(image_path))
        date_key = extract_sort_key(image_path)
        self.site_label.config(text=f"Site: {site} ({self.global_image_idx + 1}/{len(self.all_image_paths)})")
        self.date_label.config(text=f"Date: {date_key[2]}/{date_key[1]}")

        if self.selected_folder.get() != site and self.selected_folder.get() != "All Sites":
            self.folder_dropdown_programmatic_update = True
            self.selected_folder.set(site)
            self.folder_dropdown_programmatic_update = False

        p_faint = image_results.get('P_faint', 0.0)
        p_clear = image_results.get('P_clear', 0.0)
        prob_color = "green" if image_results.get('Present', False) else "red"
        self.ds_results_label.config(text=f"Certainty of Presence: {p_present:.1%} (Faint: {p_faint:.1%}) (Clear: {p_clear:.1%})", fg=prob_color)
        Tooltip(self.ds_results_label, f"P(Present) >= {Config.DS_PRESENCE_THRESHOLD*100}% or Weighted Median >= 1")

    def _update_validation_ui(self, annotations, original_image_name, p_present):
        """Updates the supervisor validation section of the UI."""
        for widget in self.validation_scrollable_frame.winfo_children(): widget.destroy()
        self.validation_vars.clear()

        validation_required = Config.DS_VALIDATION_LOWER_BOUND <= p_present <= Config.DS_VALIDATION_UPPER_BOUND
        
        columns_container = Frame(self.validation_scrollable_frame); columns_container.pack(fill='x', expand=True)

        if not validation_required and not self.force_validation_override:
            override_frame = Frame(columns_container); override_frame.pack(pady=20)
            msg = f"Validation not required (Certainty: {p_present:.1%})"
            Label(override_frame, text=msg, font=("Arial", 12, "italic"), fg="grey").pack(side='left', padx=10)
            Button(override_frame, text="Force Manual Validation", command=self.force_manual_validation, font=("Arial", 9)).pack(side='left')
        else:
            annotating_users = sorted(list(set(ann['user'] for ann in annotations)))
            num_pages = max(1, (len(annotating_users) + self.validation_users_per_page - 1) // self.validation_users_per_page)
            self.validation_current_page = min(self.validation_current_page, num_pages - 1)
            start_idx = self.validation_current_page * self.validation_users_per_page
            users_on_page = annotating_users[start_idx : start_idx + self.validation_users_per_page]

            self.page_label.config(text=f"Page {self.validation_current_page + 1} of {num_pages}")
            self.prev_page_button.config(state='normal' if self.validation_current_page > 0 else 'disabled')
            self.next_page_button.config(state='normal' if self.validation_current_page < num_pages - 1 else 'disabled')

            for user in users_on_page:
                self._create_user_validation_ui(user, columns_container, annotations)

    def _update_right_canvas(self, annotations):
        """Updates the right canvas with either a heatmap or user drawings."""
        if self.show_heatmap:
            self.full_res_right_img = self._generate_heatmap_image(self.full_res_left_img, annotations)
        else:
            self.full_res_right_img = self.full_res_left_img.copy()
            draw = ImageDraw.Draw(self.full_res_right_img)
            base_image_name = annotations[0].get('image') if annotations else None
            if base_image_name:
                all_related_annotations = [ann for ann_list in self.aggregated_data.values() for ann in ann_list if ann.get('image') == base_image_name]
                for ann in all_related_annotations:
                    if self.user_visibility.get(ann['user'], True) and ann['score'] in [1, 2] and ann['drawing']:
                        try:
                            drawing_data = json.loads(ann['drawing'])
                            line_width = 18 if ann['score'] == 2 else 8
                            user_color = self.user_color_map.get(ann['user'], '#FFFFFF')
                            for stroke in drawing_data:
                                if len(stroke) > 1: draw.line(stroke, fill=user_color, width=line_width)
                        except (json.JSONDecodeError, TypeError): continue

    def fit_and_show_image(self, canvas, pil_img):
        """Crops, resizes, and displays a PIL image based on the view_rect."""
        if pil_img is None: return
        canvas_w, canvas_h = canvas.winfo_width(), canvas.winfo_height()
        if canvas_w <= 1 or canvas_h <= 1:
            canvas.after(50, lambda: self.fit_and_show_image(canvas, pil_img)); return

        img_w, img_h = pil_img.size
        x0, y0, x1, y1 = [int(self.view_rect[i] * (img_w if i % 2 == 0 else img_h)) for i in range(4)]
        cropped_img = pil_img.crop((x0, y0, x1, y1))

        ratio = min(canvas_w / cropped_img.width, canvas_h / cropped_img.height)
        new_w, new_h = max(1, int(cropped_img.width * ratio)), max(1, int(cropped_img.height * ratio))
        resized = cropped_img.resize((new_w, new_h), Image.LANCZOS)
        
        tk_img = ImageTk.PhotoImage(resized)
        canvas.delete("all")
        canvas.create_image(canvas_w / 2, canvas_h / 2, image=tk_img, anchor="center")
        canvas.image = tk_img

    # --------------------------------------------------------------------------
    # 4. Supervisor Validation UI & Logic
    # --------------------------------------------------------------------------
    def _create_user_validation_ui(self, user, parent_frame, annotations):
        """Creates the validation UI for a single user based on their state."""
        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')

        user_column_frame = Frame(parent_frame, padx=10, pady=10); user_column_frame.pack(side='left', anchor='n', fill='y', expand=True, padx=5, pady=5)

        user_annotation = next((ann for ann in annotations if ann['user'] == user), None)
        user_score = user_annotation['score'] if user_annotation else -1
        score_text_map = {0: "No Mark", 1: "Faint Mark", 2: "Clear Mark", -1: "No Annotation"}
        
        self._create_user_header(user, user_column_frame, score_text_map[user_score])

        saved_validation = self.supervisor_validations.get(original_image_name, {}).get(user, {})
        supervisor_choice = saved_validation.get('choice')
        is_finalized = 'weight' in saved_validation

        if is_finalized:
            self._create_finalized_summary(user, user_column_frame, saved_validation)
        elif not supervisor_choice:
            self._create_initial_choice_buttons(user, user_column_frame)
        else:
            self._create_checklist_ui(user, user_column_frame, user_score, saved_validation)

    def _create_user_header(self, user, parent_frame, score_text):
        """Creates the clickable user name header with tooltip."""
        is_visible = self.user_visibility.get(user, True)
        display_color = self.user_color_map.get(user, 'black') if is_visible else 'grey'
        
        expertise = self.user_expertise_scores.get(user, 1)
        weight = self.user_weights.get(user, 1.0)
        consistency = self.user_consistency_scores.get(user, 1.0)
        tooltip_text = f"Expertise: {expertise} | Weight: {weight:.1f} | consistency: {consistency:.2f}"

        user_name_label = Label(parent_frame, text=f"{user} - {score_text}", font=("Arial", 10, "bold"), fg=display_color, cursor="hand2")
        user_name_label.pack(anchor='w', pady=(0, 5))
        user_name_label.bind("<Button-1>", lambda e, u=user: self.toggle_user_visibility(u))
        Tooltip(user_name_label, tooltip_text)

    def _create_finalized_summary(self, user, parent_frame, saved_validation):
        """Displays the summary of a completed validation."""
        final_choice = saved_validation.get('choice', 'N/A').capitalize()
        final_weight = saved_validation.get('weight', 1.0)
        final_tags = saved_validation.get('tags', [])

        summary_frame = Frame(parent_frame); summary_frame.pack(anchor='w', pady=5)
        Label(summary_frame, text=f"Choice: {final_choice}", font=("Arial", 8, "bold")).pack(anchor='w')
        if final_tags:
            Label(summary_frame, text=f"Tags: {', '.join(final_tags)}", font=("Arial", 8, "italic")).pack(anchor='w')
        Label(summary_frame, text=f"Weight: {final_weight:.3f}", font=("Arial", 8)).pack(anchor='w')
        # --- FIX: Use the passed 'user' variable directly for a reliable command ---
        Button(parent_frame, text="Reset", font=("Arial", 9), command=lambda u=user: self.reset_validation(u)).pack(pady=(10,0))

    def _create_initial_choice_buttons(self, user, parent_frame):
        """Creates the Agree/Disagree/Unsure buttons."""
        button_frame = Frame(parent_frame); button_frame.pack(anchor='w', pady=5)
        Button(button_frame, text="Agree", command=lambda u=user: self._apply_pre_validation(u, 'agree')).pack(side='left')
        Button(button_frame, text="Unsure", command=lambda u=user: self._apply_pre_validation(u, 'unsure')).pack(side='left', padx=5)
        Button(button_frame, text="Disagree", command=lambda u=user: self._apply_pre_validation(u, 'disagree')).pack(side='left')

    def _create_checklist_ui(self, user, parent_frame, user_score, saved_validation):
        """Creates the checklist UI for Agree/Disagree choices."""
        self.validation_vars[user] = {}
        saved_tags = saved_validation.get('tags', [])
        
        if saved_validation.get('choice') == 'agree':
            categories_to_show = ["Drivers"] if user_score > 0 else ["Absence Reason"]
        else: # Disagree
            categories_to_show = ["Missed Positives"] if user_score == 0 else ["False Positives"]

        for category in categories_to_show:
            factors = Config.VALIDATION_FACTORS.get(category, [])
            Label(parent_frame, text=category, font=("Arial", 9, "italic")).pack(anchor='w')
            
            if category == "Missed Positives":
                var = StringVar(value=saved_tags[0] if saved_tags else "")
                self.validation_vars[user]['missed_choice'] = var
                for factor in factors:
                    Checkbutton(parent_frame, text=factor, variable=var, onvalue=factor, offvalue="", font=("Arial", 8), command=lambda u=user: self._check_validation_constraints(u)).pack(anchor='w', padx=15)
            else:
                for factor in factors:
                    var = BooleanVar(value=(factor in saved_tags))
                    cb = Checkbutton(parent_frame, text=factor, variable=var, font=("Arial", 8), command=lambda u=user: self._check_validation_constraints(u))
                    cb.pack(anchor='w', padx=15)
                    self.validation_vars[user][factor] = var

        button_frame = Frame(parent_frame); button_frame.pack(pady=(10, 0))
        submit_button = Button(button_frame, text="Submit", font=("Arial", 9, "bold"), command=lambda u=user: self.finalize_validation(u))
        submit_button.pack(side='left', padx=5)
        self.submit_buttons[user] = submit_button
        self._check_validation_constraints(user)
        Button(button_frame, text="Reset", font=("Arial", 9), command=lambda u=user: self.reset_validation(u)).pack(side='left', padx=5)

    def _check_validation_constraints(self, user):
        """Checks checklist selections and enables/disables the Submit button."""
        submit_button = self.submit_buttons.get(user)
        if not submit_button: return

        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')
        validation_data = self.supervisor_validations.get(original_image_name, {}).get(user, {})
        supervisor_choice = validation_data.get('choice')
        
        user_annotation = next((ann for ann in self.aggregated_data.get(original_image_name, []) if ann['user'] == user), None)
        user_score = user_annotation['score'] if user_annotation else -1
        
        var_dict = self.validation_vars.get(user, {})
        is_valid = False
        if supervisor_choice == 'agree':
            is_valid = True
        elif supervisor_choice == 'disagree':
            if user_score == 0: # Missed Positives
                is_valid = bool(var_dict.get('missed_choice', StringVar()).get())
            else: # False Positives
                selected_count = sum(1 for var in var_dict.values() if isinstance(var, BooleanVar) and var.get())
                is_valid = 1 <= selected_count <= 2
        
        submit_button.config(state='normal' if is_valid else 'disabled')

    # --------------------------------------------------------------------------
    # 5. Supervisor Validation Actions
    # --------------------------------------------------------------------------
    def _apply_pre_validation(self, user, choice):
        """Saves the Agree/Disagree/Unsure choice and updates the UI."""
        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')

        if original_image_name not in self.supervisor_validations:
            self.supervisor_validations[original_image_name] = {}
        self.supervisor_validations[original_image_name][user] = {'choice': choice}

        if choice == 'unsure':
            self.finalize_validation(user)
        else:
            self.update_display(reset_view=False)

    def finalize_validation(self, user):
        """Calculates and saves the final validation based on checklist items."""
        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')
        
        validation_data = self.supervisor_validations.get(original_image_name, {}).get(user, {})
        if not validation_data: return

        supervisor_choice = validation_data.get('choice')
        user_annotation = next((ann for ann in self.aggregated_data.get(original_image_name, []) if ann['user'] == user), None)
        user_score = user_annotation['score'] if user_annotation else -1

        var_dict = self.validation_vars.get(user, {})
        weight = Config.WEIGHT_BASELINE
        selected_tags = []

        if supervisor_choice == 'agree':
            if user_score > 0: # Agree with presence, apply Driver deltas
                selected_tags = [factor for factor, var in var_dict.items() if isinstance(var, BooleanVar) and var.get()]
                weight += sum(Config.VALIDATION_DELTAS.get(tag, 0) for tag in selected_tags)
        elif supervisor_choice == 'disagree':
            if user_score == 0: # Missed Positive
                tag = var_dict.get('missed_choice', StringVar()).get()
                if tag: selected_tags = [tag]; weight += Config.VALIDATION_DELTAS.get(tag, 0)
            else: # False Positive
                selected_tags = [factor for factor, var in var_dict.items() if isinstance(var, BooleanVar) and var.get()]
                weight += sum(Config.VALIDATION_DELTAS.get(tag, 0) for tag in selected_tags)

        weight = max(Config.WEIGHT_MIN_CLAMP, min(Config.WEIGHT_MAX_CLAMP, weight))
        
        validation_data.update({'tags': selected_tags, 'score': weight, 'weight': weight})
        self.supervisor_validations[original_image_name][user] = validation_data
        
        print(f"\nFinalized Validation: Image: '{original_image_name}', User: '{user}' -> Choice: {supervisor_choice}, Final Weight: {weight:.3f}")

        try:
            with open(Config.VALIDATIONS_FILE, 'w') as f:
                json.dump(self.supervisor_validations, f, indent=2)
        except IOError as e:
            print(f"Error: Could not save validations to file: {e}")

        self._run_ds_model()
        self.update_display(reset_view=False)

    def reset_validation(self, user):
        """Resets the validation for a single user on the current image."""
        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')

        if original_image_name in self.supervisor_validations and user in self.supervisor_validations[original_image_name]:
            print(f"\nResetting validation for User: '{user}', Image: '{original_image_name}'")
            del self.supervisor_validations[original_image_name][user]
            if not self.supervisor_validations[original_image_name]:
                del self.supervisor_validations[original_image_name]

            try:
                with open(Config.VALIDATIONS_FILE, 'w') as f:
                    json.dump(self.supervisor_validations, f, indent=2)
                print("Validations file updated.")
            except IOError as e:
                print(f"Error: Could not save validations to file: {e}")

            self._run_ds_model()
            self.update_display(reset_view=False)
        else:
            print("No validation to reset for this user.")

    def force_manual_validation(self):
        """Clears existing validations, sets a flag to override the certainty lock, and redisplays the UI."""
        print("\nOverriding certainty lock to allow manual validation.")
        image_path = self.all_image_paths[self.global_image_idx]
        original_image_name = os.path.basename(image_path).replace('_qc', '')
        if original_image_name in self.supervisor_validations:
            print(f"Clearing existing validations for '{original_image_name}' due to override.")
            del self.supervisor_validations[original_image_name]
        
        self.force_validation_override = True
        self.update_display(reset_view=False)

    # --------------------------------------------------------------------------
    # 6. GUI Event Handlers
    # --------------------------------------------------------------------------
    def toggle_user_visibility(self, user):
        """Toggles the visibility of a user's drawings."""
        if user in self.user_visibility:
            self.user_visibility[user] = not self.user_visibility[user]
            print(f"Toggled visibility for {user}: {'Visible' if self.user_visibility[user] else 'Hidden'}")
            self.update_display(reset_view=False)

    def toggle_heatmap(self):
        """Toggles the heatmap view on the right canvas."""
        self.show_heatmap = not self.show_heatmap
        self.update_display(reset_view=False)

    def on_scroll_folders(self, event):
        """Allows changing the folder dropdown by scrolling."""
        current_values = self.folder_dropdown['values']
        try:
            current_index = current_values.index(self.selected_folder.get())
        except ValueError:
            current_index = 0
        
        delta = -1 if event.delta > 0 else 1
        new_index = max(0, min(len(current_values) - 1, current_index + delta))
            
        if new_index != current_index:
            self.selected_folder.set(current_values[new_index])
            self.on_folder_select()
        return "break"

    def on_folder_select(self, *args):
        """Filters the image list when a folder is selected from the dropdown."""
        if self.folder_dropdown_programmatic_update: return
        folder = self.selected_folder.get()
        
        if folder == "All Sites":
            self.global_image_idx = 0
        else:
            try:
                self.global_image_idx = next(i for i, path in enumerate(self.all_image_paths) if os.path.basename(os.path.dirname(path)) == folder)
            except StopIteration:
                self.global_image_idx = 0
        self.update_display()

    def prev_validation_page(self):
        """Navigates to the previous page of user checklists."""
        if self.validation_current_page > 0:
            self.validation_current_page -= 1
            self.update_display(reset_view=False)

    def next_validation_page(self):
        """Navigates to the next page of user checklists."""
        self.validation_current_page += 1
        self.update_display(reset_view=False)

    # --------------------------------------------------------------------------
    # 7. Image Navigation & View Control
    # --------------------------------------------------------------------------
    def next_image(self):
        """Navigates to the next image in the list."""
        if self.global_image_idx < len(self.all_image_paths) - 1:
            self.global_image_idx += 1
            self.update_display(reset_view=True)

    def prev_image(self):
        """Navigates to the previous image in the list."""
        if self.global_image_idx > 0:
            self.global_image_idx -= 1
            self.update_display(reset_view=True)

    def start_pan(self, event):
        """Records the starting point for a pan operation."""
        self.pan_start_x, self.pan_start_y = event.x, event.y
        self.pan_start_view_rect = list(self.view_rect)
        event.widget.config(cursor="fleur")

    def pan_image(self, event):
        """Pans the image by updating the view_rect."""
        if self.pan_start_view_rect is None or self.full_res_left_img is None: return
        
        view_w, view_h = self.pan_start_view_rect[2] - self.pan_start_view_rect[0], self.pan_start_view_rect[3] - self.pan_start_view_rect[1]
        delta_x_pixels, delta_y_pixels = event.x - self.pan_start_x, event.y - self.pan_start_y
        
        img_w, img_h = self.full_res_left_img.size
        scale = (view_w * img_w) / event.widget.winfo_width() if event.widget.winfo_width() > 0 else 1
        
        delta_x = (delta_x_pixels * scale) / img_w
        delta_y = (delta_y_pixels * scale) / img_h # Note: y-scale is based on width scale for consistent panning

        new_x0 = np.clip(self.pan_start_view_rect[0] - delta_x, 0, 1 - view_w)
        new_y0 = np.clip(self.pan_start_view_rect[1] - delta_y, 0, 1 - view_h)
        self.view_rect = [new_x0, new_y0, new_x0 + view_w, new_y0 + view_h]
        
        self.update_display(reset_view=False)

    def end_pan(self, event):
        """Ends the pan operation."""
        self.pan_start_view_rect = None
        event.widget.config(cursor="")

    def zoom(self, event):
        """Zooms in or out of the image, centered on the mouse cursor."""
        zoom_factor = 1.1 if event.delta > 0 else 1 / 1.1
        
        if self.zoom_level * zoom_factor < 1:
            self.zoom_level, self.view_rect = 1.0, [0, 0, 1, 1]
        else:
            self.zoom_level *= zoom_factor
            mouse_x = event.x / event.widget.winfo_width()
            mouse_y = event.y / event.widget.winfo_height()

            view_w, view_h = self.view_rect[2] - self.view_rect[0], self.view_rect[3] - self.view_rect[1]
            img_x, img_y = self.view_rect[0] + mouse_x * view_w, self.view_rect[1] + mouse_y * view_h

            new_view_w, new_view_h = 1 / self.zoom_level, (1 / self.zoom_level) * (self.full_res_left_img.height / self.full_res_left_img.width)

            new_x0 = np.clip(img_x - mouse_x * new_view_w, 0, 1 - new_view_w)
            new_y0 = np.clip(img_y - mouse_y * new_view_h, 0, 1 - new_view_h)
            self.view_rect = [new_x0, new_y0, new_x0 + new_view_w, new_y0 + new_view_h]
            
        self.update_display(reset_view=False)

    # --------------------------------------------------------------------------
    # 8. Image Generation
    # --------------------------------------------------------------------------
    def _generate_heatmap_image(self, base_image, annotations):
        """Generates a transparent heatmap overlay and composites it onto the base image."""
        if base_image is None: return None
        base_pil_img = base_image.copy().convert("RGBA")
        h, w = base_pil_img.height, base_pil_img.width
        heatmap = np.zeros((h, w), dtype=np.float32)

        for ann in annotations:
            if ann['score'] in [1, 2] and ann['drawing']:
                try:
                    drawing_data = json.loads(ann['drawing'])
                    intensity = 255 if ann['score'] == 2 else 128
                    line_width = 18 if ann['score'] == 2 else 8
                    for stroke in drawing_data:
                        if len(stroke) > 1:
                            pts = np.array(stroke, dtype=np.int32).reshape((-1, 1, 2))
                            cv2.polylines(heatmap, [pts], isClosed=False, color=(intensity), thickness=line_width)
                except (json.JSONDecodeError, TypeError): continue

        if np.any(heatmap > 0):
            heatmap = cv2.GaussianBlur(heatmap, (99, 99), 0)
        
        heatmap_normalized = cv2.normalize(heatmap, None, 0, 255, cv2.NORM_MINMAX, dtype=cv2.CV_8U)
        heatmap_colored_bgr = cv2.applyColorMap(heatmap_normalized, cv2.COLORMAP_JET)
        heatmap_colored_rgba = cv2.cvtColor(heatmap_colored_bgr, cv2.COLOR_BGR2RGBA)
        heatmap_colored_rgba[:, :, 3] = heatmap_normalized

        heatmap_pil = Image.fromarray(heatmap_colored_rgba)
        blended_img = Image.alpha_composite(base_pil_img, heatmap_pil)
        return blended_img.convert("RGB")

if __name__ == "__main__":
    root = Tk()
    app = SupervisorApp(root)
    root.mainloop()